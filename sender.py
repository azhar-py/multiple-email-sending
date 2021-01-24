#By Azhar Khan
import smtplib
import openpyxl
import time
from email.message import EmailMessage
from pyfiglet import Figlet
f =Figlet(font='5lineoblique')
print(f.renderText("Email Sender"))

Email = input("Enter Email Address")
password = input("Enter password")

count = 0
if Email.find("gmail.com") == -1:
    print("Enter Valied Gmail Account ! ")
    print('Hope to see you again! Goodbye!')
    exit()
else:
    try:
        with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
            smtp.login(Email, password)
            print("Login Account \n ..........\n.........\n........ ")
            print("Mailing Server starting.....")
    except:
        print("Login Fail ! 'Hope to see you again! Goodbye!' ")
        exit()
while True:
    subject = input(" Enter the Subject")
    Message = input("Write Message ")
    if subject=="":
        print("You not Enter the Subject \n press Y for Continue ")
        sb = input()
        if sb =='y' or sb =='Y':
            pass
    if Message=="":
        conform = input('Message is Empty \n' "Press Y for Continue")
        if conform == 'Y' or conform=='y':
            break
    else:
        break

obj = openpyxl.load_workbook('Emaillist.xlsx')
sheet =obj.active
row =sheet.max_row
col = sheet.max_column
for r in range(1,row+1):
    for c in range(1,col+1):
        to = sheet.cell(row=r,column=c).value

    if to == None:
        continue
    try:
        msg = EmailMessage()
        to_mail = to
        print(to_mail)
        msg['Subject'] = subject
        msg['From'] = Email
        msg['To'] = to_mail
        msg.set_content(Message)
        with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
            smtp.login(Email, password)
            smtp.send_message(msg)
            print("Mail Send to " + to_mail + "\n")
            count = count+1
            if count == row:
                print("Sending the All Mail")
                print("This Script Writen by Azhar khan ")
                exit()
            else:
                time.sleep(5)
    except:
        print("Fail Sending Mail " + to_mail )

